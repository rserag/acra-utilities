# importing required classes
import re
import json
import argparse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers, NamedStyle, Font
from pypdf import PdfReader

# Function to parse the details string into a JSON-compatible dictionary
def parse_details(details):
    # Split the details into lines and parse key-value pairs
    lines = details.split("\n")
    result = {}
    for line in lines:
        # Match key-value pairs with either `:` or `-` as delimiters
        match = re.match(r"(.+?)[:`-]\s*(.+)", line.strip())
        if match:
            key, value = match.groups()
            result[key.strip()] = value.strip()
    return result

# Function to print logs in case debbuging is enabled
def debug_log(message):
    if args.debug:
        print(f"[DEBUG] {message}")

# Create the parser
parser = argparse.ArgumentParser(description="A simple program to process arguments")

# Define command-line arguments
parser.add_argument('-n', '--name', type=str, help='File name', required=True)
parser.add_argument('--debug', action='store_true', help='Enable debug messages')

# Parse the arguments
args = parser.parse_args()

# creating a pdf reader object
reader = PdfReader(args.name)

# printing number of pages in pdf file
debug_log(f"Number of pages in the PDF: {len(reader.pages)}")

# Extract text from all pages in a PDF and concatenate them
text = ""
for page in reader.pages:
    text += page.extract_text()

# Regular expression to capture blocks between "1  ՎԱՐԿ" and "Նշումներ գրավի վերաբերյալ"
pattern = r"(\d+\s+ՎԱՐԿ)(.*?)Նշումներ գրավի վերաբերյալ"
matches = re.findall(pattern, text, re.DOTALL)

# Create a list of extracted objects
results = []
for match in matches:
    # Parse the details string
    details_json = parse_details(match[1].strip())

    # Each match[0] is the number (e.g., "1  ՎԱՐԿ"), match[1] is the detailed content
    results.append({
        "loan_number": match[0].strip(),
        "details": details_json
    })

# Convert the final object to JSON
final_json = json.dumps(results, ensure_ascii=False, indent=4)

# Print the JSON object
debug_log(final_json)

# Create a new Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Loans"

# Create a custom style for AMD currency
amd_currency_style = NamedStyle(name="AMD_Currency")
amd_currency_style.number_format = '[$֏-hy-AM]#,##0'

# Write the headers
headers = ["Loan Number"] + list(results[0]["details"].keys())
for col_num, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_num, value=header)

# Write the data rows
for row_num, entry in enumerate(results, start=2):
    # Extract only the numeric part of the loan_number using regex
    loan_number = re.search(r'\d+', entry["loan_number"]).group()
    ws.cell(row=row_num, column=1, value=loan_number)  # Write the loan number

    # Write the details
    for col_num, key in enumerate(entry["details"].keys(), start=2):
        value = entry["details"][key]
        if isinstance(value, str) and "AMD" in value:  # Format fields with "AMD"
            value = value.replace("AMD", "").replace(",", "").strip()  # Remove AMD suffix and commas
            ws.cell(row=row_num, column=col_num, value=int(value))  # Set as a number
            ws.cell(row=row_num, column=col_num).style = amd_currency_style  # Apply currency format
        else:
            ws.cell(row=row_num, column=col_num, value=value)  # Write other values as-is

# Add "Totals" row at the end
total_row = len(results) + 2
ws.cell(row=total_row, column=1, value="Totals")
ws.cell(row=total_row, column=1).font = Font(bold=True)

# Calculate totals for numeric columns
for col_num in range(2, len(headers) + 1):  # Start from column 2 (skipping Loan Number)
    col_letter = get_column_letter(col_num)
    try:
        # Check if the column has numeric values
        col_values = [
            ws.cell(row=row_num, column=col_num).value
            for row_num in range(2, total_row)
            if isinstance(ws.cell(row=row_num, column=col_num).value, (int, float))
        ]
        if col_values:  # If numeric values exist in the column
            ws.cell(row=total_row, column=col_num, value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})")
            ws.cell(row=total_row, column=col_num).style = amd_currency_style
    except Exception as e:
        pass

# Adjust column widths
for col_num, header in enumerate(headers, start=1):
    ws.column_dimensions[get_column_letter(col_num)].width = max(15, len(header) + 2)

# Save the workbook to a file named args.name
wb.save("{0}.xlsx".format(args.name))

print("XLSX file '{0}.xlsx' has been created successfully!".format(args.name))
